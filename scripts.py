from sap_popup_utils import sap_with_popup_guard
import win32com.client
import os
import time
import random
import string
from datetime import datetime, timedelta
from robot_framework.initialize_sap import initialize_sap
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
from pathlib import Path
def SDStamdataTabel(orchestrator_connection = None):
    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    application = SapGuiAuto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    with sap_with_popup_guard():
        # Opret SAP-objekter (som i VBS)
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        connection = application.Children(0)
        session = connection.Children(0)

        # Trin som i VBS
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "ke5x"
        session.findById("wnd[0]").sendVKey(0)

        session.findById("wnd[0]/usr/ctxtGD_PCGRP").text = "2"
        session.findById("wnd[0]/usr/ctxtGD_PCGRP").setFocus()
        session.findById("wnd[0]/usr/ctxtGD_PCGRP").caretPosition = 1
        session.findById("wnd[0]/tbar[1]/btn[8]").press()  # Execute

        # Højreklik i ALV og vælg "&XXL" (Export -> Spreadsheet)
        grid = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell")
        grid.contextMenu()
        grid.selectContextMenuItem("&XXL")

        # Vent kort på formatdialog (wnd[1])
        def _exists(id_):
            try:
                session.findById(id_)
                return True
            except Exception:
                return False

        deadline = time.time() + 8
        while time.time() < deadline and not _exists("wnd[1]/usr"):
            time.sleep(0.1)

        # Vælg "Andre formater" + XLSX (key="10")
        usr = session.findById("wnd[1]/usr")
        usr.findById("radRB_OTHERS").setFocus()
        usr.findById("radRB_OTHERS").select()
        usr.findById("cmbG_LISTBOX").setFocus()
        usr.findById("cmbG_LISTBOX").Key = "10"  # Excel (Office 2007 XLSX-format)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()  # OK

        # Vent på gem-dialogfelter (kan være samme wnd[1])
        deadline = time.time() + 8
        while time.time() < deadline and not (_exists("wnd[1]/usr/ctxtDY_PATH") and _exists("wnd[1]/usr/ctxtDY_FILENAME")):
            time.sleep(0.1)

        excel_before = _pids()
        # Sti og filnavn (hardcoded som i VBS)
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = os.getcwd()
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "Stamdatatabel.XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len("Stamdatatabel.XLSX")
        session.findById("wnd[1]/tbar[0]/btn[11]").press() 

        # Tilbage/luk (to gange)
        try:
            session.findById("wnd[0]/tbar[0]/btn[3]").press()
        except Exception:
            pass
        try:
            session.findById("wnd[0]/tbar[0]/btn[3]").press()
        except Exception:
            pass


    # Luk eventuelle NYE Excel-processer (åbner ofte forsinket)
    new_pids = close_new_excels(excel_before, wait_seconds=30)
    print(f"Lukkede Excel PIDs: {sorted(new_pids)}")

    # Luk ALT SAP (alle sessions/forbindelser)
    close_all_sap()


def SDForfaldneFaktura(orchestrator_connection = None):
    user = orchestrator_connection.get_constant("SapUsernameRPA").value
    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    application = SapGuiAuto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    with sap_with_popup_guard():
        # ========== 1) Navigér til transaktionen ==========
        session.findById("wnd[0]").resizeWorkingPane(209, 30, False)
        session.findById("wnd[0]/tbar[0]/okcd").text = "zmir6n"
        session.findById("wnd[0]").sendVKey(0)

        # Vælg variant / bruger
        session.findById("wnd[0]/tbar[1]/btn[17]").press()
        session.findById("wnd[1]/usr/txtENAME-LOW").text = user
        session.findById("wnd[1]/usr/txtENAME-LOW").setFocus()
        session.findById("wnd[1]/usr/txtENAME-LOW").caretPosition = len(user)
        session.findById("wnd[1]/tbar[0]/btn[8]").press()

        # Vælg række i ALV (samme som VB: current cell + selected row = "2")
        alv = session.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell")
        alv.setCurrentCell(2, "TEXT")
        alv.selectedRows = "2"
        session.findById("wnd[1]/tbar[0]/btn[2]").press()  # Vælg

        # Kør rapport
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]").sendVKey(0)

        # Højreklik i ALV og vælg "Export → Spreadsheet" (&XXL)
        grid = session.findById("wnd[0]/usr/cntlCUSTOM_CONTROL/shellcont/shell")
        grid.setCurrentCell(10, "GJAHR")
        grid.contextMenu()
        grid.selectContextMenuItem("&XXL")

        # ========== 2) Formateringsdialog: vælg XLSX ==========
        # (venter kort hvis dialogen ikke er fremme helt endnu)
        def _exists(id_):
            try:
                session.findById(id_)
                return True
            except:
                return False

        deadline = time.time() + 8
        while time.time() < deadline and not _exists("wnd[1]/usr"):
            time.sleep(0.1)

        # Vælg "Vælg fra alle disponible formater" + dropdown key=10
        usr = session.findById("wnd[1]/usr")
        usr.findById("radRB_OTHERS").select()
        usr.findById("cmbG_LISTBOX").Key = "10"  # Excel (Office 2007 XLSX-format)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()  # OK

        # ========== 3) Gem-dialog: sti og filnavn ==========
        # (nogle systemer viser den på samme vindue; ellers ny popup)
        deadline = time.time() + 8
        while time.time() < deadline and not (_exists("wnd[1]/usr/ctxtDY_PATH") and _exists("wnd[1]/usr/ctxtDY_FILENAME")):
            time.sleep(0.1)

        excel_before = _pids()

        session.findById("wnd[1]/usr/ctxtDY_PATH").text = os.getcwd()
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "Forfaldne fakturaer MTM.XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len("Forfaldne fakturaer MTM.XLSX")
        session.findById("wnd[1]/tbar[0]/btn[11]").press()  # Gem

        # ========== 4) Ryd op (tilbage/luk) ==========
        try:
            session.findById("wnd[0]/tbar[0]/btn[15]").press()
        except:
            pass
        try:
            session.findById("wnd[0]/tbar[0]/btn[15]").press()
        except:
            pass
      
    # Luk eventuelle NYE Excel-processer (åbner ofte forsinket)
    new_pids = close_new_excels(excel_before, wait_seconds=30)
    print(f"Lukkede Excel PIDs: {sorted(new_pids)}")

    # Luk ALT SAP (alle sessions/forbindelser)
    close_all_sap()

def _pids(process_name="EXCEL.EXE"):
    try:
        out = subprocess.check_output(
            ['tasklist', '/FI', f'IMAGENAME eq {process_name}', '/FO', 'CSV', '/NH'],
            text=True, creationflags=0x08000000
        )
    except Exception:
        return set()
    pids = set()
    for line in out.splitlines():
        if not line.strip():
            continue
        parts = [s.strip('"') for s in line.split(',')]
        if len(parts) >= 2 and parts[0].upper() == process_name.upper():
            try:
                pids.add(int(parts[1]))
            except ValueError:
                pass
    return pids

def close_new_excels(before_pids, wait_seconds=30):
    """Vent op til wait_seconds på nye Excel-processer og luk kun dem."""
    deadline = time.time() + wait_seconds
    new = set()
    while time.time() < deadline:
        now = _pids()
        new = now - before_pids
        if new:
            break
        time.sleep(0.5)
    for pid in new:
        os.system(f'taskkill /PID {pid} /F >NUL 2>&1')
    return new

def close_all_sap():
    """Luk alle SAP sessions/connections og til sidst processen."""
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        app = SapGuiAuto.GetScriptingEngine
    except Exception:
        app = None

    if app:
        try:
            for ci in range(app.Children.Count):
                conn = app.Children(ci)
                for si in range(conn.Children.Count - 1, -1, -1):
                    try:
                        sess = conn.Children(si)
                        # prøv pænt logoff
                        try:
                            sess.findById("wnd[0]/tbar[0]/okcd").text = "/nex"
                            sess.findById("wnd[0]").sendVKey(0)
                            time.sleep(0.5)
                        except Exception:
                            pass
                        # og luk vinduet hvis stadig åbent
                        try:
                            sess.findById("wnd[0]/tbar[0]/btn[15]").press()
                        except Exception:
                            pass
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            app.Quit()
        except Exception:
            pass

    # sikkerhedsnet hvis noget hænger
    os.system('taskkill /IM saplogon.exe /F >NUL 2>&1')
    os.system('taskkill /IM saplgpad.exe /F >NUL 2>&1')

def SDLonUdtrak(orchestrator_connection = None):

    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    application = SapGuiAuto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    with sap_with_popup_guard():
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "cji3"
        session.findById("wnd[0]").sendVKey(0)

        session.findById("wnd[0]/tbar[1]/btn[17]").press()
        session.findById("wnd[1]/usr/txtV-LOW").text = "SD Udtræk MTM"
        session.findById("wnd[1]/usr/txtENAME-LOW").text = ""
        session.findById("wnd[1]/usr/txtENAME-LOW").setFocus()
        session.findById("wnd[1]/usr/txtENAME-LOW").caretPosition = 0
        session.findById("wnd[1]/tbar[0]/btn[8]").press()

        session.findById("wnd[0]/tbar[1]/btn[21]").press()
        shell = session.findById("wnd[0]/shellcont/shellcont/shell/shellcont[0]/shell")
        shell.expandNode("          1")
        shell.selectNode("         23")
        shell.topNode = "          1"
        shell.doubleClickNode("         23")

        today = datetime.today()
        fifteen_days_ago = today - timedelta(days=15)
        date_format = lambda d: d.strftime("%d.%m.%Y")

        session.findById("wnd[0]/usr/ctxt%%DYN001-LOW").text = date_format(fifteen_days_ago)
        session.findById("wnd[0]/usr/ctxt%%DYN001-HIGH").text = date_format(today)
        session.findById("wnd[0]/usr/ctxt%%DYN001-HIGH").setFocus()
        session.findById("wnd[0]/usr/ctxt%%DYN001-HIGH").caretPosition = 10
        session.findById("wnd[0]/tbar[0]/btn[11]").press()

        session.findById("wnd[0]/tbar[1]/btn[8]").press()

        # Tag baseline for Excel-PIDs LIGE inden eksport/filskrivning initieres
        excel_before = _pids()

        session.findById("wnd[0]/tbar[1]/btn[43]").press()

        # --- TJEK/VALG AF FORMAT + EVT. STED/ FILNAVN ---
        def _exists(id_):
            try:
                session.findById(id_)
                return True
            except Exception:
                return False

        deadline = time.time() + 8
        while time.time() < deadline and not _exists("wnd[1]"):
            time.sleep(0.1)

        if _exists("wnd[1]"):
            is_format = _exists("wnd[1]/usr/radRB_OTHERS") and _exists("wnd[1]/usr/cmbG_LISTBOX")
            if is_format:
                usr = session.findById("wnd[1]/usr")
                usr.findById("radRB_OTHERS").select()
                usr.findById("cmbG_LISTBOX").Key = "10"  # din XLSX-key
                # lad popuppen komme igen fremover (hvis felt findes)
                try:
                    usr.findById("chkCB_ALWAYS").Selected = False
                except Exception:
                    pass

                has_save_fields = _exists("wnd[1]/usr/ctxtDY_PATH") and _exists("wnd[1]/usr/ctxtDY_FILENAME")
                if not has_save_fields:
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()  # OK
                    deadline = time.time() + 8
                    while time.time() < deadline and not (_exists("wnd[1]/usr/ctxtDY_PATH") and _exists("wnd[1]/usr/ctxtDY_FILENAME")):
                        time.sleep(0.1)

            if _exists("wnd[1]/usr/ctxtDY_PATH") and _exists("wnd[1]/usr/ctxtDY_FILENAME"):
                session.findById("wnd[1]/usr/ctxtDY_PATH").text = os.getcwd()
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "export.xlsx"
                session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 11
                # her trykkes den faktiske GEM/OK – Excel kan poppe op EFTER dette
                session.findById("wnd[1]/tbar[0]/btn[11]").press()  # (brug btn[0] hvis det er jeres "OK")
            else:
                print("Ingen gem-felter vist; SAP gemmer til standardplacering.")

        else:
            print("Format-popup blev ikke vist; SAP bruger husket format.")

        # Luk undermenuer/dialoger i hovedvinduet
        try:
            session.findById("wnd[0]/tbar[0]/btn[15]").press()
        except Exception:
            pass
        try:
            session.findById("wnd[0]/tbar[0]/btn[15]").press()
        except Exception:
            pass

    # <-- UDENFOR sap_with_popup_guard(): nu rydder vi op, så watcher ikke hænger

    # Luk eventuelle NYE Excel-processer (åbner ofte forsinket)
    new_pids = close_new_excels(excel_before, wait_seconds=30)
    print(f"Lukkede Excel PIDs: {sorted(new_pids)}")

    # Luk ALT SAP (alle sessions/forbindelser)
    close_all_sap()
    
def InputToTemplate():
    """
    Kopiér skabelon -> læs export.xlsx!Sheet1 fra A2 -> skriv til dagsdatokopi ved A2
    -> refresh på faner ('Tabel1', 'Antal linjer pr. navn') uden baggrundsrefresh
    -> gem, luk og frigiv fil-låse deterministisk.
    Returnerer Path til den nye fil (eller None).
    """
    from pathlib import Path
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries, get_column_letter
    import win32com.client as win32
    import win32com, importlib, shutil, os, time
    import win32process

    print("[InputToTemplate] Start")

    BASE = Path.cwd()
    EXPORT    = BASE / "export.xlsx"
    TEMPLATE  = BASE / "SDLoen, indtastningsskabelon.xlsx"
    NAME = f"{datetime.today().strftime('%d.%m.%Y')}SDLoen.xlsx"
    OUTFILE   = BASE / NAME
    READ_SHEET     = "Sheet1"
    TARGET_SHEET   = "Sheet1"
    REFRESH_SHEETS = ["Tabel1", "Antal linjer pr. navn"]
    START_ROW, START_COL = 2, 1  # A2

    # ---------- helpers ----------
    def _get_excel_app():
        print("[Excel] Starter Excel.Application (late-binding)")
        try:
            return win32.DispatchEx("Excel.Application")
        except Exception as e1:
            print(f"[Excel] DispatchEx fejlede: {e1}. Rydder gencache og prøver igen...")
            try:
                win32.gencache.is_readonly = False
                gen_path = win32.gencache.GetGeneratePath()
                shutil.rmtree(gen_path, ignore_errors=True)
                shutil.rmtree(os.path.join(os.path.dirname(win32com.__file__), "gen_py"), ignore_errors=True)
                importlib.reload(win32.gencache)
            except Exception as e2:
                print(f"[Excel] Kunne ikke rydde gencache: {e2}")
            return win32.DispatchEx("Excel.Application")

    def _excel_pid(xl_app):
        try:
            hwnd = xl_app.Hwnd
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
            return pid
        except Exception:
            return None

    def _disable_background_refresh(wb):
        print("[Excel] Slår baggrundsrefresh fra på Connections/QueryTables/ListObjects")
        # Connections
        try:
            for conn in wb.Connections:
                try:
                    if hasattr(conn, "ODBCConnection"):
                        conn.ODBCConnection.BackgroundQuery = False
                except Exception:
                    pass
                try:
                    if hasattr(conn, "OLEDBConnection"):
                        conn.OLEDBConnection.BackgroundQuery = False
                except Exception:
                    pass
        except Exception:
            pass
        # QueryTables / ListObjects
        try:
            for ws in wb.Worksheets:
                try:
                    for lo in ws.ListObjects:
                        try:
                            qt = lo.QueryTable
                            if qt:
                                qt.BackgroundQuery = False
                        except Exception:
                            pass
                except Exception:
                    pass
                try:
                    for qt in ws.QueryTables:
                        try:
                            qt.BackgroundQuery = False
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass

    def _wait_refresh_done(wb, timeout=180):
        print(f"[Excel] Venter på at refresh bliver færdig (timeout={timeout}s)...")
        t0 = time.time()
        while time.time() - t0 < timeout:
            busy = False
            # Connections
            try:
                for conn in wb.Connections:
                    try:
                        if conn.Refreshing:
                            busy = True
                            break
                    except Exception:
                        pass
            except Exception:
                pass
            # QueryTables
            if not busy:
                try:
                    for ws in wb.Worksheets:
                        try:
                            for lo in ws.ListObjects:
                                try:
                                    qt = lo.QueryTable
                                    if qt and qt.Refreshing:
                                        busy = True; break
                                except Exception:
                                    pass
                            if busy: break
                        except Exception:
                            pass
                        try:
                            for qt in ws.QueryTables:
                                try:
                                    if qt.Refreshing:
                                        busy = True; break
                                except Exception:
                                    pass
                            if busy: break
                        except Exception:
                            pass
                except Exception:
                    pass
            if not busy:
                print("[Excel] Refresh er færdig.")
                return True
            time.sleep(0.5)
        print("[Excel] Refresh timeout – fortsætter.")
        return False

    def _kill_pid(pid):
        if pid:
            print(f"[Excel] KILL PID {pid}")
            try:
                os.system(f'taskkill /PID {pid} /F >NUL 2>&1')
            except Exception:
                pass

    def _last_used_from(ws, min_row=2, min_col=1):
        # hurtig afgrænsning af dataområde fra og med (min_row, min_col)
        dim = ws.calculate_dimension()  # fx 'A1:F1234'
        min_c, min_r, max_c, max_r = range_boundaries(dim)
        sr, sc = max(min_row, min_r), max(min_col, min_c)
        last_row = sr - 1
        last_col = sc - 1
        if max_r < sr or max_c < sc:
            return last_row, last_col
        for r_idx, row in enumerate(ws.iter_rows(min_row=sr, max_row=max_r,
                                                min_col=sc, max_col=max_c,
                                                values_only=True), start=sr):
            row_last_col = None
            for j in range(len(row) - 1, -1, -1):
                v = row[j]
                if v not in (None, ""):
                    row_last_col = sc + j
                    break
            if row_last_col is not None:
                last_row = r_idx
                if row_last_col > last_col:
                    last_col = row_last_col
        return last_row, last_col

    # ---------- 1) læs data hurtigt fra export.xlsx ----------
    print(f"[Read] Åbner {EXPORT}")
    if not EXPORT.exists():
        print("[Read] export.xlsx findes ikke.")
        return None

    wb_src = load_workbook(EXPORT, data_only=True, read_only=True)
    ws_src = wb_src[READ_SHEET]
    last_row, last_col = _last_used_from(ws_src, min_row=START_ROW, min_col=START_COL)
    print(f"[Read] Sidste brugte område fra A2: last_row={last_row}, last_col={last_col}")

    if last_row >= START_ROW and last_col >= START_COL:
        data_rows = list(ws_src.iter_rows(min_row=START_ROW, max_row=last_row,
                                          min_col=START_COL, max_col=last_col,
                                          values_only=True))
        data_rows = [tuple("" if v is None else v for v in row) for row in data_rows]
    else:
        data_rows = []
    wb_src.close()

    if not data_rows:
        print("[Read] Ingen data fundet fra A2 og ned. Stopper.")
        return None

    rows, cols = len(data_rows), len(data_rows[0])
    print(f"[Read] Læste {rows} rækker x {cols} kolonner")

    # ---------- 2) kopi af skabelon ----------
    print(f"[Copy] Kopierer skabelon: {TEMPLATE} -> {OUTFILE}")
    if not TEMPLATE.exists():
        print("[Copy] Skabelon findes ikke.")
        return None
    try:
        shutil.copyfile(TEMPLATE, OUTFILE)
    except Exception as e:
        print(f"[Copy] Fejl ved kopiering: {e}")
        return None

    # ---------- 3) skriv + refresh + gem + luk ----------
    xl = _get_excel_app()
    xl.Visible = False
    xl.DisplayAlerts = False
    pid = _excel_pid(xl)
    print(f"[Excel] PID={pid}")

    prev_calc = None
    try:
        print(f"[Excel] Åbner OUTFILE: {OUTFILE}")
        wb = xl.Workbooks.Open(Filename=str(OUTFILE))

        # performance (best effort) EFTER åbning
        try:
            prev_calc = xl.Calculation
            print(f"[Excel] Current CalcMode={prev_calc}")
        except Exception:
            pass
        for attr, val in (("ScreenUpdating", False), ("EnableEvents", False)):
            try:
                setattr(xl, attr, val)
                print(f"[Excel] {attr}={val}")
            except Exception:
                pass
        try:
            xl.Calculation = -4135  # xlCalculationManual
            print("[Excel] Calculation=Manual")
        except Exception:
            print("[Excel] Kunne ikke sætte Calculation=Manual (fortsætter)")

        # skriv blokken i ét skud
        print(f"[Write] Skriver til {TARGET_SHEET}!A2 ...")
        try:
            ws = wb.Worksheets(TARGET_SHEET)
        except Exception:
            names = [sh.Name for sh in wb.Worksheets]
            raise RuntimeError(f"Fanen '{TARGET_SHEET}' findes ikke. Tilgængelige: {names}")

        start_col_letter = get_column_letter(START_COL)
        end_col_letter   = get_column_letter(START_COL + cols - 1)
        end_row          = START_ROW + rows - 1
        rng_addr = f"{start_col_letter}{START_ROW}:{end_col_letter}{end_row}"
        print(f"[Write] Range={rng_addr}")

        max_cols = max(len(r) for r in data_rows)
        if max_cols != cols:
            cols = max_cols
        data_rect = tuple(tuple("" if i >= len(r) or r[i] is None else r[i] for i in range(cols))
                          for r in data_rows)

        ws.Range(rng_addr).Value2 = data_rect
        wb.Save()
        print("[Write] Data skrevet og gemt")

        # slå baggrundsrefresh fra og refresh de relevante faner
        _disable_background_refresh(wb)

        def refresh_sheet(name):
            print(f"[Refresh] Aktiverer '{name}' og kører RefreshAll")
            try:
                wb.Worksheets(name).Activate()
            except Exception as e:
                print(f"[Refresh] Kunne ikke aktivere '{name}': {e}")
                return
            wb.RefreshAll()
            done = _wait_refresh_done(wb, timeout=180)
            if not done:
                print("[Refresh] Timeout – fallback sleep(120)")
                time.sleep(120)
            wb.Save()
            print(f"[Refresh] '{name}' færdig og gemt")

        for sheet_name in REFRESH_SHEETS:
            refresh_sheet(sheet_name)

        wb.Close(SaveChanges=True)
        print("[Excel] Workbook lukket")

    except Exception as e:
        print(f"[Excel] Fejl: {e}")
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            pass
        raise
    finally:
        # gendan og quit
        try:
            if prev_calc is not None:
                xl.Calculation = prev_calc
                print(f"[Excel] Gendanner CalcMode={prev_calc}")
        except Exception:
            pass
        for attr, val in (("EnableEvents", True), ("ScreenUpdating", True)):
            try:
                setattr(xl, attr, val)
            except Exception:
                pass
        try:
            xl.Quit()
            print("[Excel] Quit() kaldt")
        except Exception:
            print("[Excel] Quit() fejlede")

        # sikkerhedsnet – kill hvis Excel stadig hænger
        _kill_pid(pid)
        time.sleep(0.5)  # giv OS/OneDrive tid til at slippe låse

    print(f"[Done] Skrev {rows} x {cols} til {OUTFILE}")
    return OUTFILE, NAME
